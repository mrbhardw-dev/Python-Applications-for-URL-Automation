# SCRIPT TO FORMAT REPORT

import openpyxl, os
from .fn_message import click_file_error
from icecream import ic


def format_report():
    my_location = os.getcwd()
    reports_location = f'{my_location}\\REPORT\\'
    excel_file = f'{reports_location}app-report.xlsx'

    workbook = openpyxl.load_workbook(excel_file)
    print('Loading workbok', workbook)

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        print('Formating sheet:', ws)

        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 4)
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.auto_filter.ref = ws.dimensions

    workbook.save(f'{reports_location}app-report.xlsx')

    workbook.close()